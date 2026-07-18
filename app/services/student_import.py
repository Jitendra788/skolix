"""Parse CSV / Excel student roster imports and build StudentCreate payloads."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from typing import Any

from dateutil import parser as date_parser
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel

from .. import schemas

_MAX_IMPORT_BYTES = 5 * 1024 * 1024

# Canonical column keys (first row headers; aliases normalized the same way).
HEADER_ALIASES: dict[str, frozenset[str]] = {
    "admission_no": frozenset(
        {"admission_no", "registration_no", "reg_no", "registration_number"}
    ),
    "full_name": frozenset(
        {"full_name", "student_name", "name", "student_full_name"}
    ),
    "class_name": frozenset({"class_name", "class", "grade"}),
    "date_of_admission": frozenset(
        {"date_of_admission", "admission_date", "doi"}
    ),
    "section": frozenset({"section", "sec"}),
    "parent_phone": frozenset(
        {"parent_phone", "sms_phone", "mobile_for_sms", "phone", "mobile"}
    ),
    "parent_name": frozenset({"parent_name", "guardian_name"}),
    "date_of_birth": frozenset({"date_of_birth", "dob", "birth_date"}),
    "gender": frozenset({"gender", "sex"}),
    "discount_fee_percent": frozenset(
        {"discount_fee_percent", "discount_pct", "discount", "fee_discount_pct"}
    ),
    "birth_form_nic": frozenset({"birth_form_nic", "nic", "student_nic", "bform"}),
    "orphan": frozenset({"orphan"}),
    "caste": frozenset({"caste"}),
    "osc": frozenset({"osc"}),
    "identification_mark": frozenset(
        {"identification_mark", "id_mark", "identifying_mark"}
    ),
    "previous_school": frozenset({"previous_school", "last_school"}),
    "religion": frozenset({"religion"}),
    "blood_group": frozenset({"blood_group", "blood"}),
    "previous_board_roll": frozenset(
        {"previous_board_roll", "board_roll", "previous_roll"}
    ),
    "family": frozenset({"family"}),
    "disease": frozenset({"disease", "medical"}),
    "additional_note": frozenset({"additional_note", "notes", "remarks"}),
    "total_siblings": frozenset({"total_siblings", "siblings"}),
    "address": frozenset({"address", "home_address"}),
    "father_name": frozenset({"father_name", "father"}),
    "father_national_id": frozenset({"father_national_id", "father_cnic"}),
    "father_occupation": frozenset({"father_occupation"}),
    "father_education": frozenset({"father_education"}),
    "father_mobile": frozenset({"father_mobile", "father_phone"}),
    "father_profession": frozenset({"father_profession"}),
    "father_income": frozenset({"father_income"}),
    "mother_name": frozenset({"mother_name", "mother"}),
    "mother_national_id": frozenset({"mother_national_id", "mother_cnic"}),
    "mother_occupation": frozenset({"mother_occupation"}),
    "mother_education": frozenset({"mother_education"}),
    "mother_mobile": frozenset({"mother_mobile", "mother_phone"}),
    "mother_profession": frozenset({"mother_profession"}),
    "mother_income": frozenset({"mother_income"}),
}

EXTRAS_KEYS = frozenset(
    {
        "date_of_admission",
        "discount_fee_percent",
        "birth_form_nic",
        "orphan",
        "caste",
        "osc",
        "identification_mark",
        "previous_school",
        "religion",
        "blood_group",
        "previous_board_roll",
        "family",
        "disease",
        "additional_note",
        "total_siblings",
        "address",
        "father_name",
        "father_national_id",
        "father_occupation",
        "father_education",
        "father_mobile",
        "father_profession",
        "father_income",
        "mother_name",
        "mother_national_id",
        "mother_occupation",
        "mother_education",
        "mother_mobile",
        "mother_profession",
        "mother_income",
    }
)

_TEMPLATE_HEADERS: list[str] = [
    "admission_no",
    "full_name",
    "class_name",
    "date_of_admission",
    "section",
    "parent_phone",
    "parent_name",
    "date_of_birth",
    "gender",
    "discount_fee_percent",
    "birth_form_nic",
    "orphan",
    "caste",
    "osc",
    "identification_mark",
    "previous_school",
    "religion",
    "blood_group",
    "previous_board_roll",
    "family",
    "disease",
    "additional_note",
    "total_siblings",
    "address",
    "father_name",
    "father_national_id",
    "father_occupation",
    "father_education",
    "father_mobile",
    "father_profession",
    "father_income",
    "mother_name",
    "mother_national_id",
    "mother_occupation",
    "mother_education",
    "mother_mobile",
    "mother_profession",
    "mother_income",
]

_TEMPLATE_SAMPLE_ROW: list[str] = [
    "STU-SAMPLE-001",
    "Sample Student",
    "Class 1",
    "2026-04-01",
    "A",
    "+92 300 0000000",
    "Father Name & Mother Name",
    "2015-05-10",
    "Male",
    "",
    "",
    "No",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "",
    "123 Sample Street",
    "Father Name",
    "",
    "",
    "",
    "+92 300 1111111",
    "",
    "",
    "Mother Name",
    "",
    "",
    "",
    "+92 300 2222222",
    "",
    "",
]


def norm_header(s: str) -> str:
    t = (s or "").strip().replace("\ufeff", "").lower()
    t = re.sub(r"[\s\-\./]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t


def _alias_map() -> dict[str, str]:
    m: dict[str, str] = {}
    for canon, aliases in HEADER_ALIASES.items():
        m[norm_header(canon)] = canon
        for a in aliases:
            m[norm_header(a)] = canon
    return m


ALIAS_TO_CANONICAL = _alias_map()


def _parse_date_value(val: Any) -> tuple[str | None, str | None]:
    if val is None:
        return None, None
    if isinstance(val, datetime):
        return val.date().isoformat(), None
    if isinstance(val, date):
        return val.isoformat(), None
    if isinstance(val, (int, float)):
        try:
            d = from_excel(val)
            if isinstance(d, datetime):
                return d.date().isoformat(), None
            if isinstance(d, date):
                return d.isoformat(), None
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return None, None
    try:
        dt = date_parser.parse(s, dayfirst=False, yearfirst=True)
        return dt.date().isoformat(), None
    except Exception:
        return None, f"unrecognized date: {s!r}"


def _row_values_to_canonical(
    headers: list[str], values: list[Any]
) -> dict[str, str]:
    out: dict[str, str] = {}
    for h, v in zip(headers, values):
        key = norm_header(h)
        if not key:
            continue
        canon = ALIAS_TO_CANONICAL.get(key)
        if not canon:
            continue
        if v is None:
            s = ""
        elif isinstance(v, float) and canon in (
            "date_of_admission",
            "date_of_birth",
        ):
            iso, _ = _parse_date_value(v)
            s = iso or str(v).strip()
        else:
            s = str(v).strip() if v is not None else ""
        if s == "":
            continue
        out[canon] = s
    return out


def parse_csv_bytes(raw: bytes) -> tuple[list[tuple[int, dict[str, str]]], str | None]:
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        return [], "File must be UTF-8 encoded CSV."
    rdr = csv.reader(io.StringIO(text))
    rows = list(rdr)
    if not rows:
        return [], "The file is empty."
    headers = rows[0]
    data_rows: list[tuple[int, dict[str, str]]] = []
    for i, cells in enumerate(rows[1:], start=2):
        if not any((c or "").strip() for c in cells):
            continue
        canon = _row_values_to_canonical(headers, cells)
        if canon:
            data_rows.append((i, canon))
    if not data_rows:
        return [], "No data rows found (only a header or blank lines)."
    return data_rows, None


def parse_xlsx_bytes(raw: bytes) -> tuple[list[tuple[int, dict[str, str]]], str | None]:
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as e:
        return [], f"Could not read Excel file: {e}"
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], "The spreadsheet is empty."
        headers = [str(c) if c is not None else "" for c in header_row]
        data_rows: list[tuple[int, dict[str, str]]] = []
        row_num = 2
        for cells in rows_iter:
            if not cells or not any(
                c is not None and str(c).strip() != "" for c in cells
            ):
                row_num += 1
                continue
            values = list(cells)
            if len(values) < len(headers):
                values.extend([None] * (len(headers) - len(values)))
            canon: dict[str, str] = {}
            for hi, h in enumerate(headers):
                key = norm_header(h)
                if not key:
                    continue
                ckey = ALIAS_TO_CANONICAL.get(key)
                if not ckey:
                    continue
                v = values[hi] if hi < len(values) else None
                if v is None or (isinstance(v, str) and not v.strip()):
                    continue
                if ckey in ("date_of_admission", "date_of_birth"):
                    iso, err = _parse_date_value(v)
                    if err:
                        canon[ckey] = str(v).strip()
                    elif iso:
                        canon[ckey] = iso
                else:
                    canon[ckey] = str(v).strip()
            if canon:
                data_rows.append((row_num, canon))
            row_num += 1
        if not data_rows:
            return [], "No data rows found under the header row."
        return data_rows, None
    finally:
        wb.close()


def parse_import_file(
    raw: bytes, filename: str
) -> tuple[list[tuple[int, dict[str, str]]], str | None]:
    if len(raw) > _MAX_IMPORT_BYTES:
        return [], f"File too large (max {_MAX_IMPORT_BYTES // (1024 * 1024)} MB)."
    low = (filename or "").lower()
    if low.endswith(".csv"):
        return parse_csv_bytes(raw)
    if low.endswith(".xlsx"):
        return parse_xlsx_bytes(raw)
    return [], "Unsupported file type. Use .csv or .xlsx."


def build_student_create_from_row(
    row: dict[str, str],
    valid_class_names: set[str],
) -> tuple[schemas.StudentCreate | None, str | None]:
    admission_no = (row.get("admission_no") or "").strip()
    full_name = (row.get("full_name") or "").strip()
    class_name = (row.get("class_name") or "").strip()
    date_adm = (row.get("date_of_admission") or "").strip()

    if not full_name:
        return None, "Missing required field: full_name."
    if not admission_no:
        return None, "Missing required field: admission_no."
    if not class_name:
        return None, "Missing required field: class_name."
    if class_name not in valid_class_names:
        return None, (
            f"Unknown class_name {class_name!r}. It must match a class from "
            "School setup exactly (check spelling and spacing)."
        )
    if not date_adm:
        return None, "Missing required field: date_of_admission."

    iso_adm, err = _parse_date_value(date_adm)
    if err or not iso_adm:
        return None, f"Invalid date_of_admission: {err or date_adm!r}."

    dob_raw = (row.get("date_of_birth") or "").strip()
    dob_iso = ""
    if dob_raw:
        d_iso, d_err = _parse_date_value(dob_raw)
        if d_err or not d_iso:
            return None, f"Invalid date_of_birth: {d_err or dob_raw!r}."
        dob_iso = d_iso

    section = (row.get("section") or "").strip()
    sms = (row.get("parent_phone") or "").strip()
    parent_name_in = (row.get("parent_name") or "").strip()
    gender = (row.get("gender") or "").strip()

    father_name = (row.get("father_name") or "").strip()
    mother_name = (row.get("mother_name") or "").strip()
    father_mobile = (row.get("father_mobile") or "").strip()
    mother_mobile = (row.get("mother_mobile") or "").strip()

    parent_phone = sms or father_mobile or mother_mobile
    if parent_name_in:
        parent_name = parent_name_in
    else:
        parts = [father_name, mother_name]
        parent_name = " & ".join([p for p in parts if p])

    extras: dict[str, str] = {"date_of_admission": iso_adm}
    for k in EXTRAS_KEYS:
        if k == "date_of_admission":
            continue
        v = (row.get(k) or "").strip()
        if v:
            extras[k] = v

    item = schemas.StudentCreate(
        admission_no=admission_no,
        full_name=full_name,
        class_name=class_name,
        section=section,
        parent_phone=parent_phone,
        parent_name=parent_name,
        date_of_birth=dob_iso,
        gender=gender,
        admission_extras=extras,
    )
    return item, None


def template_csv_bytes() -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow(_TEMPLATE_HEADERS)
    w.writerow(_TEMPLATE_SAMPLE_ROW)
    return buf.getvalue().encode("utf-8-sig")


def template_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(_TEMPLATE_HEADERS)
    ws.append(_TEMPLATE_SAMPLE_ROW)
    ws2 = wb.create_sheet("How_to_fill")
    ws2["A1"] = (
        "Required columns: admission_no, full_name, class_name, date_of_admission."
    )
    ws2["A2"] = (
        "date_of_admission and date_of_birth: use YYYY-MM-DD or a standard date; "
        "Excel date cells are accepted."
    )
    ws2["A3"] = (
        "class_name must match your School setup exactly (e.g. Class 1). "
        "Delete the sample row and add one row per student."
    )
    ws2["A4"] = (
        "parent_phone is used for SMS/WhatsApp; if empty, father_mobile or "
        "mother_mobile is used."
    )
    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()
